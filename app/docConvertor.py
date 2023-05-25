from openpyxl import *
import io
from tempfile import NamedTemporaryFile
from openpyxl.utils import get_column_letter
def create_xls(result):
    wb = Workbook()
    ws = wb[wb.sheetnames[0]]
    ws.title = "Отчет об анализе"
    ws.cell(row=1, column=1, value="Адрес объекта")
    ws.cell(row=1, column=2, value="Необходимые работы")
    ws.cell(row=1, column=3, value="Статус срочности работ по адресу")
    column_widths = []
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    currentrow = 2

    for address_data in result['result']:
        ws.cell(currentrow, 1, value= address_data['adress'])
        ws.merge_cells(start_row=currentrow, start_column=1, end_row=currentrow + len(address_data['workname']) - 1, end_column=1)
        for (work_count, work_data) in enumerate(address_data['workname']):
            ws.cell(row=currentrow + work_count, column=2, value=work_data)
        currentrow += len(address_data['workname'])
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        return io.BytesIO(stream)

def create_xlsx(result):
    wb = Workbook()
    ws = wb[wb.sheetnames[0]]
    ws.title = "Отчет об анализе"
    ws.cell(row=1, column=1, value="Адрес объекта")
    ws.cell(row=1, column=2, value="Необходимые работы")
    ws.cell(row=1, column=3, value="Статус срочности работ по адресу")
    column_widths = []
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    currentrow = 2

    for address_data in result['result']:
        ws.cell(currentrow, 1, value= address_data['adress'])
        ws.merge_cells(start_row=currentrow, start_column=1, end_row=currentrow + len(address_data['workname']) - 1, end_column=1)
        for (work_count, work_data) in enumerate(address_data['workname']):
            ws.cell(row=currentrow + work_count, column=2, value=work_data)
        currentrow += len(address_data['workname'])
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        return io.BytesIO(stream)